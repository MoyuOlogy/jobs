#!/usr/bin/env python3
"""
清华就业信息网招聘信息获取工具
爬取清华大学学生职业发展指导中心招聘信息，支持列表 + 详情页全文。

用法:
  python3 qinghuajobs.py                        # 查看前 3 页招聘信息
  python3 qinghuajobs.py --pages 5              # 指定翻页数
  python3 qinghuajobs.py --date 2026-05-15      # 筛选指定日期
  python3 qinghuajobs.py --today                # 筛选当天的
  python3 qinghuajobs.py --detail               # 带详情（爬每个职位的详情页）
  python3 qinghuajobs.py --save                 # 保存 Excel 到 output/tsinghua/
  python3 qinghuajobs.py --detail --save        # 详情 + 保存
"""

import json
import os
import re
import sys
from datetime import datetime, timezone, timedelta

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ── 路径 ──────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")
OUTPUT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "output", "tsinghua"))
TZ = timezone(timedelta(hours=8))
BASE = "https://career.cic.tsinghua.edu.cn"

# ── 配置 ──────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


# ── 列表 API ──────────────────────────────────────────
def fetch_list(pgno=1):
    """获取招聘信息列表页"""
    resp = requests.post(
        f"{BASE}/xsglxt/b/jyxt/anony/xxfbForWx",
        data={"type": "ZPXX", "pgno": pgno},
        headers={
            "User-Agent": "Mozilla/5.0",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        timeout=15,
    )
    return resp.json()


# ── 详情页爬取 ────────────────────────────────────────
def fetch_detail(zpxxid):
    """爬取单个职位的详情页，返回提取的文本"""
    url = f"{BASE}/xsglxt/f/jyxt/anony/showZwxxForWx?zpxxid={zpxxid}&type=ZPXX"
    try:
        resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
        html = resp.text

        # 取 #xxfb 区域，如果没有就取整个 body
        body = re.search(r'id="xxfb"[^>]*>(.*?)</div>', html, re.DOTALL)
        if not body:
            body = re.search(r'<body[^>]*>(.*?)</body>', html, re.DOTALL)

        text = ""
        if body:
            text = body.group(1)
            text = re.sub(r'<script[^>]*>.*?</script>', "", text, flags=re.DOTALL)
            text = re.sub(r'<style[^>]*>.*?</style>', "", text, flags=re.DOTALL)
            text = re.sub(r'<br\s*/?>', "\n", text)
            text = re.sub(r'<[^>]+>', "", text)
            text = re.sub(r'\n{3,}', "\n\n", text)
            text = re.sub(r'&nbsp;', " ", text)
            text = text.strip()

        # 提取字段
        xl = ""
        m = re.search(r"学历[：:]\s*([^\n]+)", text)
        if m:
            xl = m.group(1).strip()

        zy = ""
        m = re.search(r"专业[：:]\s*([^\n]+)", text)
        if m:
            zy = m.group(1).strip()

        email = ""
        for p in [r"邮箱[：:]\s*([^\s]+)", r"E-mail[：:]\s*([^\s]+)", r"@"]:
            m = re.search(p, text)
            if m:
                raw = text
                if "@" in raw:
                    email_m = re.search(r"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", raw)
                    if email_m:
                        email = email_m.group(1)
                break

        return {
            "text": text,
            "summary": text[:2000] if text else "(无正文)",
            "xl": xl or "见正文",
            "zy": zy or "见正文",
            "email": email,
        }
    except Exception as e:
        return {"text": "", "summary": f"❌ 抓取失败: {e}", "xl": "", "zy": "", "email": ""}


# ── 数据处理 ──────────────────────────────────────────
def get_detail_url(zpxxid):
    return f"{BASE}/xsglxt/f/jyxt/anony/showZwxxForWx?zpxxid={zpxxid}&type=ZPXX"


def filter_by_date(items, date_str):
    """按发布日期筛选"""
    return [i for i in items if i.get("fbsj") == date_str]


# ── Excel 输出 ────────────────────────────────────────
STYLE_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
STYLE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
STYLE_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
STYLE_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_excel(items, has_detail, filepath):
    """写入 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "招聘信息"

    if has_detail:
        headers = ["序号", "职位名称", "单位名称", "工作地点", "单位性质",
                    "发布日期", "学历要求", "专业要求", "联系邮箱", "职位摘要", "详情链接"]
    else:
        headers = ["序号", "职位名称", "单位名称", "工作地点", "单位性质",
                    "发布日期", "是否置顶", "详情链接"]

    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = STYLE_HEADER_FONT
        cell.fill = STYLE_HEADER_FILL
        cell.alignment = STYLE_HEADER_ALIGN
        cell.border = STYLE_BORDER

    for i, item in enumerate(items, 1):
        url = get_detail_url(item["id"])
        if has_detail:
            d = item.get("_detail", {})
            ws.append([
                i, item["zwmc"], item["dwmc"], item["gzdqmc"],
                item["dwxz"], item["fbsj"],
                d.get("xl", ""), d.get("zy", ""), d.get("email", ""),
                d.get("summary", ""), url,
            ])
        else:
            ws.append([
                i, item["zwmc"], item["dwmc"], item["gzdqmc"],
                item["dwxz"], item["fbsj"],
                "⭐" if item.get("sfzd") == "是" else "",
                url,
            ])

    # 列宽
    if has_detail:
        widths = [6, 40, 30, 18, 12, 12, 12, 22, 28, 60, 50]
    else:
        widths = [6, 40, 30, 18, 12, 12, 10, 50]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath


# ── CLI ───────────────────────────────────────────────
def main():
    cfg = load_config()
    default_pages = cfg.get("default_pages", 3)

    # 解析参数
    pages = default_pages
    date_filter = None
    need_detail = False
    need_save = False

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--pages" and i + 1 < len(args):
            pages = int(args[i + 1])
            i += 2
        elif args[i] == "--date" and i + 1 < len(args):
            date_filter = args[i + 1]
            i += 2
        elif args[i] == "--today":
            date_filter = datetime.now(TZ).strftime("%Y-%m-%d")
            i += 1
        elif args[i] == "--detail":
            need_detail = True
            i += 1
        elif args[i] == "--save":
            need_save = True
            i += 1
        else:
            print(f"❌ 未知参数: {args[i]}")
            return

    # ── 拉取列表 ──
    print(f"📄 正在拉取前 {pages} 页...")
    all_items = []
    for pg in range(1, pages + 1):
        try:
            data = fetch_list(pg)
            all_items.extend(data)
            print(f"  第 {pg} 页: {len(data)} 条", end="")
            print(" ✅" if data else " ❌")
        except Exception as e:
            print(f"  第 {pg} 页: ❌ {e}")

    # 筛选日期
    if date_filter:
        all_items = filter_by_date(all_items, date_filter)
        print(f"\n📅 筛选 {date_filter}: {len(all_items)} 条")

    print(f"\n📋 共 {len(all_items)} 条招聘信息\n")

    # ── 拉取详情 ──
    if need_detail and all_items:
        print("🔍 正在爬取详情页...")
        for idx, item in enumerate(all_items, 1):
            print(f"  [{idx}/{len(all_items)}] {item['zwmc'][:25]}...", end=" ")
            item["_detail"] = fetch_detail(item["id"])
            print(f"✅ {len(item['_detail']['summary'])} 字")

    # ── 显示 ──
    if not need_save:
        show_basic(all_items, need_detail)
        return

    # ── 保存 Excel ──
    today_str = datetime.now(TZ).strftime("%m%d")
    label = date_filter.replace("-", "") if date_filter else today_str
    filename = f"清华就业信息网_{label}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    write_excel(all_items, need_detail, filepath)
    mode = "含详情" if need_detail else "基础"
    print(f"\n✅ 已保存 ({mode}): {filepath}")


def show_basic(items, has_detail):
    """打印到终端"""
    for i, item in enumerate(items[:10], 1):  # 只显示前 10
        d = item.get("_detail", {})
        print(f"{i}. {item['zwmc']}")
        print(f"   单位：{item['dwmc']} | {item['gzdqmc']} | {item['dwxz']} | {item['fbsj']}")
        if has_detail:
            print(f"   学历：{d.get('xl','')} | 专业：{d.get('zy','')} | 邮箱：{d.get('email','')}")
        print()
    if len(items) > 10:
        print(f"  ... 还有 {len(items) - 10} 条")


if __name__ == "__main__":
    main()
